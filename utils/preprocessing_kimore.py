import os
import csv
import openpyxl
import pandas as pd
import numpy as np
import joblib
from sklearn.model_selection import StratifiedKFold, LeaveOneOut
from tqdm import tqdm

index_Spine_Base=0
index_Spine_Mid=4
index_Neck=8
index_Head=12   # no orientation
index_Shoulder_Left=16
index_Elbow_Left=20
index_Wrist_Left=24
index_Hand_Left=28
index_Shoulder_Right=32
index_Elbow_Right=36
index_Wrist_Right=40
index_Hand_Right=44
index_Hip_Left=48
index_Knee_Left=52
index_Ankle_Left=56
index_Foot_Left=60  # no orientation    
index_Hip_Right=64
index_Knee_Right=68
index_Ankle_Right=72
index_Foot_Right=76   # no orientation
index_Spine_Shoulder=80
index_Tip_Left=84     # no orientation
index_Thumb_Left=88   # no orientation
index_Tip_Right=92    # no orientation
index_Thumb_Right=96  # no orientation

body_parts = [index_Spine_Base, index_Spine_Mid, index_Neck, index_Head, index_Shoulder_Left, index_Elbow_Left, index_Wrist_Left, index_Hand_Left, index_Shoulder_Right, index_Elbow_Right, index_Wrist_Right, index_Hand_Right, index_Hip_Left, index_Knee_Left, index_Ankle_Left, index_Foot_Left, index_Hip_Right, index_Knee_Right, index_Ankle_Right, index_Ankle_Right, index_Spine_Shoulder, index_Tip_Left, index_Thumb_Left, index_Tip_Right, index_Thumb_Right
]

def convert_joints(original, num_joints=25, num_channel=3, body_part=body_parts):
    X_train = np.zeros((original.shape[0], num_joints * num_channel)).astype('float32')
    for row in range(original.shape[0]):
        counter = 0
        for parts in body_part:
            for i in range(num_channel):
                X_train[row, counter+i] = original[row, parts+i]
            counter += num_channel 
    return X_train.astype('float32')

def get_raw_data():
    path = "data/kimore"
    kinect_joints = ["spinebase", "spinemid", "neck", "head", 
                    "shoulderleft", "elbowleft", "wristleft", 
                    "handleft", "shoulderright", "elbowright", 
                    "wristright", "handright", "hipleft", "kneeleft", 
                    "ankleleft", "footleft", "hipright", "kneeright", 
                    "ankleright", "footright", "spineshoulder", "handtipleft", 
                    "thumbleft", "handtipright", "thumbright"]

    enable_kinect_joints = False
    enable_slice_list = False

    data = []
    for (root, dirs, files) in os.walk(path):
        #code from where??
        # if current directory contains "Raw", extract data
        if "Raw" in dirs:
            
            new_dict = {}
            
            # get exercise number
            new_dict["Exercise"] = int(root[-1])

            # extract raw data
            raw_files = os.listdir(os.path.join(root, "Raw"))
            for file in raw_files:

                file_path = os.path.join(os.path.join(root, "Raw"),file)
                if os.path.isdir(file_path):
                    continue
                csv_file = open(file_path, newline='')
                csv_reader = csv.reader(csv_file)
                
                if file.startswith("JointOrientation"):
                    
                    if enable_kinect_joints:
                        for joint in kinect_joints:
                            new_dict[joint + "-o"] = []

                        for row in csv_reader:
                            for i in range(len(kinect_joints)):
                                if len(row) > 0:
                                    new_dict[kinect_joints[i] + "-o"].append(row[(4*i):(4*i+4)])
                    else:
                        new_dict["JointOrientation"] = []
                        for row in csv_reader:
                            if len(new_dict["JointOrientation"]) >= 182 and enable_slice_list:
                                break
                            elif len(row) > 0:
                                if '' in row:
                                    row.remove('')
                                new_dict["JointOrientation"].append(np.array([float(i) for i in row]))
                        #np.array(new_dict["JointOrientation"])
                                

                elif file.startswith("JointPosition"):
                    
                    if enable_kinect_joints:
                        for joint in kinect_joints:
                            new_dict[joint + "-p"] = []

                        for row in csv_reader:
                            for i in range(len(kinect_joints)):
                                if len(row) > 0:
                                    new_dict[kinect_joints[i] + "-p"].append(row[(4*i):(4*i+3)])
                    else:
                        new_dict["JointPosition"] = []
                        for row in csv_reader:
                            if len(new_dict["JointPosition"]) >= 182 and enable_slice_list:
                                break
                            elif len(row) > 0:
                                if '' in row:
                                    row.remove('')
                                new_dict["JointPosition"].append(np.array([float(i) for i in row]))
                        #np.array(new_dict["JointPosition"])

                elif file.startswith("TimeStamp"):

                    new_dict["Timestamps"] = []
                    for row in csv_reader:
                        if len(new_dict["Timestamps"]) >= 182 and enable_slice_list:
                            break
                        elif len(row) > 0:
                            if '' in row:
                                    row.remove('')
                            new_dict["Timestamps"].append(row)
                                
                            

            # extract data labels
            label_files = os.listdir(os.path.join(root, "Label"))
            for file in label_files:

                file_path = os.path.join(os.path.join(root, "Label"),file)
                book = openpyxl.load_workbook(file_path)
                sheet = book.active

                if file.startswith("SuppInfo"):
                    for i in range(1, sheet.max_column):
                        t = sheet.cell(1, i).value
                        v = sheet.cell(2, i).value
                        new_dict[t] = v
                
                elif file.startswith("ClinicalAssessment"):
                    new_dict["cTS"] = sheet.cell(2, new_dict["Exercise"]+1).value
                    new_dict["cPO"] = sheet.cell(2, new_dict["Exercise"]+6).value
                    new_dict["cCF"] = sheet.cell(2, new_dict["Exercise"]+11).value
            
            # append exercise to data
            data.append(new_dict)
    df = pd.DataFrame(data)
    df=df.dropna()
    df=df.reset_index(drop=True)
    dfnew=df.dropna().copy()
    exs={}
    for i in range(len(dfnew)):
        if sum([j.shape[0]==100 for j in dfnew.iloc[i]["JointPosition"]])!=len(dfnew.iloc[i]["JointPosition"]):
            print("missmatched data, skipping...")
            print(dfnew.iloc[i]["Subject ID"])
            continue
        if dfnew.iloc[i]["Exercise"] not in exs:
            exs[dfnew.iloc[i]["Exercise"]]=[]
        data=np.stack(dfnew.iloc[i]["JointPosition"])
        label=dfnew.iloc[i]["cTS"]
        crex=dfnew.iloc[i]["Exercise"]
        sbid=dfnew.iloc[i]["Subject ID"]
        exs[crex].append((data, label, sbid)) #데이터, 라벨, 환자 ID 형태로 exs에 저장
    joblib.dump(exs, "./data/kimore_raw.pkl")



class Data_Loader():
    def __init__(self, path=None, loocv=False):
        self.path = path
        self.body_part = self.body_parts()       
        self.num_timestep = 100
        self.num_joints = len(self.body_part)
        #self.scaled_x, self.scaled_y = self.preprocessing()
        self.exs = joblib.load(path)
        self.loocv = loocv

    def body_parts(self):
        body_parts = [index_Spine_Base, index_Spine_Mid, index_Neck, index_Head, index_Shoulder_Left, index_Elbow_Left, index_Wrist_Left, index_Hand_Left, index_Shoulder_Right, index_Elbow_Right, index_Wrist_Right, index_Hand_Right, index_Hip_Left, index_Knee_Left, index_Ankle_Left, index_Foot_Left, index_Hip_Right, index_Knee_Right, index_Ankle_Right, index_Ankle_Right, index_Spine_Shoulder, index_Tip_Left, index_Thumb_Left, index_Tip_Right, index_Thumb_Right]
        return body_parts

    def convert_joints(self, original, num_joints=25, num_channel=3):
        """
        Convert the original data to a 2D array with shape (num_samples, num_joints * num_channel)
        """
        body_part=self.body_parts()
        X_train = np.zeros((original.shape[0], num_joints * num_channel)).astype('float32')
        for row in range(original.shape[0]):
            counter = 0
            for parts in body_part:
                for i in range(num_channel):
                    X_train[row, counter+i] = original[row, parts+i]
                counter += num_channel 
        return X_train.astype('float32')
    
    def getsubjects(self, exs, exercise):
        subjects = [i[2] for i in exs[exercise]]
        return list(set(subjects))
    
    @staticmethod
    def ispatient(st):
        #환자는 앞에 BPS가 ID에 붙어있습니다.
        if st[0] in ['B', 'P', 'S']:
            return True
        else:
            return False
    
    def select_k(self, subjects, k, random_state=0):
        #subject에서 k-fold로 나누는 함수
        #selected_subjects[i]와 train_subjects[i]는 같은 fold에 속하는 subject들입니다.
        stf=StratifiedKFold(n_splits=k, shuffle=True, random_state=random_state) if not self.loocv else LeaveOneOut()
        selected_subjects = []
        train_subjects = []
        iter_split=stf.split(subjects, [self.ispatient(i) for i in subjects]) if not self.loocv else stf.split(subjects)
        for train, test in iter_split:
            selected_subjects.append([subjects[i] for i in test])
            train_subjects.append([subjects[i] for i in train])
        return selected_subjects, train_subjects
    
    def getdata(self, exs, exercise, subjects):
        #exs에서 exercise에 해당하는 subject들의 데이터들을 반환합니다.
        data = []
        labels = []
        temporal = int(100)
        for i in exs[exercise]:
            if i[2] in subjects:
                mydata=self.convert_joints(i[0])
                #mydata is (time, joints*3)                
                timelimit=(mydata.shape[0]//temporal)*temporal
                mydata=mydata[:timelimit]
                mydata=mydata.reshape((-1, temporal, 25, 3))
                #mydata is (n, 100, joints*3)
                for j in range(mydata.shape[0]):
                    data.append(mydata[j])
                    labels.append(i[1])
        return np.stack(data), np.array(labels)
    
    def getkdata(self, exercise, k=None):
        #k-fold로 나누어진 train, test data를 반환합니다.
        sbs=self.getsubjects(self.exs, exercise)
        selected_subjects, train_subjects = self.select_k(sbs, k)
        train_data = []
        train_labels = []
        test_data = []
        test_labels = []
        for i in selected_subjects:
            data, labels = self.getdata(self.exs, exercise, i)
            test_data.append(data)
            test_labels.append(labels)
        for i in train_subjects:
            data, labels = self.getdata(self.exs, exercise, i)
            train_data.append(data)
            train_labels.append(labels)

        return train_data, train_labels, test_data, test_labels, selected_subjects, train_subjects

class PreNormalize3D:
    """PreNormalize for NTURGB+D 3D keypoints (x, y, z). Codes adapted from https://github.com/lshiwjx/2s-AGCN. """

    def unit_vector(self, vector):
        """Returns the unit vector of the vector. """
        return vector / np.linalg.norm(vector)

    def angle_between(self, v1, v2):
        """Returns the angle in radians between vectors 'v1' and 'v2'. """
        if np.abs(v1).sum() < 1e-6 or np.abs(v2).sum() < 1e-6:
            return 0
        v1_u = self.unit_vector(v1)
        v2_u = self.unit_vector(v2)
        return np.arccos(np.clip(np.dot(v1_u, v2_u), -1.0, 1.0))

    def rotation_matrix(self, axis, theta):
        """Return the rotation matrix associated with counterclockwise rotation
        about the given axis by theta radians."""
        if np.abs(axis).sum() < 1e-6 or np.abs(theta) < 1e-6:
            return np.eye(3)
        axis = np.asarray(axis)
        axis = axis / np.sqrt(np.dot(axis, axis))
        a = np.cos(theta / 2.0)
        b, c, d = -axis * np.sin(theta / 2.0)
        aa, bb, cc, dd = a * a, b * b, c * c, d * d
        bc, ad, ac, ab, bd, cd = b * c, a * d, a * c, a * b, b * d, c * d
        return np.array([[aa + bb - cc - dd, 2 * (bc + ad), 2 * (bd - ac)],
                        [2 * (bc - ad), aa + cc - bb - dd, 2 * (cd + ab)],
                        [2 * (bd + ac), 2 * (cd - ab), aa + dd - bb - cc]])

    def __init__(self, zaxis=[0, 1], xaxis=[8, 4], align_spine=True, align_center=True):
        self.zaxis = zaxis
        self.xaxis = xaxis
        self.align_spine = align_spine
        self.align_center = align_center

    def __call__(self, results):
        skeleton = results['keypoint']
        total_frames = results.get('total_frames', skeleton.shape[1])

        M, T, V, C = skeleton.shape
        assert T == total_frames
        if skeleton.sum() == 0:
            return results


        '''
        index0 = [i for i in range(T) if not np.all(np.isclose(skeleton[0, i], 0))]

        assert M in [1, 2]
        if M == 2:
            index1 = [i for i in range(T) if not np.all(np.isclose(skeleton[1, i], 0))]
            if len(index0) < len(index1):
                skeleton = skeleton[:, np.array(index1)]
                skeleton = skeleton[[1, 0]]
            else:
                skeleton = skeleton[:, np.array(index0)]
        else:
            skeleton = skeleton[:, np.array(index0)]
        '''
        T_new = skeleton.shape[1]

        if self.align_center:
            if skeleton.shape[2] == 25:
                main_body_center = skeleton[0, 0, 1].copy()
            else:
                main_body_center = skeleton[0, 0, -1].copy()
            mask = ((skeleton != 0).sum(-1) > 0)[..., None]
            skeleton = (skeleton - main_body_center) * mask

        if self.align_spine:
            joint_bottom = skeleton[0, 0, self.zaxis[0]]
            joint_top = skeleton[0, 0, self.zaxis[1]]
            axis = np.cross(joint_top - joint_bottom, [0, 0, 1])
            angle = self.angle_between(joint_top - joint_bottom, [0, 0, 1])
            matrix_z = self.rotation_matrix(axis, angle)
            skeleton = np.einsum('abcd,kd->abck', skeleton, matrix_z)

            joint_rshoulder = skeleton[0, 0, self.xaxis[0]]
            joint_lshoulder = skeleton[0, 0, self.xaxis[1]]
            axis = np.cross(joint_rshoulder - joint_lshoulder, [1, 0, 0])
            angle = self.angle_between(joint_rshoulder - joint_lshoulder, [1, 0, 0])
            matrix_x = self.rotation_matrix(axis, angle)
            skeleton = np.einsum('abcd,kd->abck', skeleton, matrix_x)

        results['keypoint'] = skeleton
        results['total_frames'] = T_new
        results['body_center'] = main_body_center
        return results

def normalize(data, preprocessor):
    #input : data: NTVC
    normalized_data = []
    for i in tqdm(range(data.shape[0])):
        processed = preprocessor({'keypoint': data[i:i+1]})
        normalized_data.append(processed['keypoint'])
    print([nd.shape for nd in normalized_data][:10])
    return normalized_data

def normalize_kfold(data, preprocessor):
    #input: 5개의 data, each NTVC
    normalized_data = []
    for fold in range(len(data)):
        fold_data = data[fold]
        norm_fold_data = []
        for i in tqdm(range(fold_data.shape[0])):
            processed = preprocessor({'keypoint': fold_data[i:i+1]})
            norm_fold_data.append(processed['keypoint'])
        normalized_data.append(np.concatenate(norm_fold_data, axis=0))
        print(normalized_data[fold].shape)
    return normalized_data

def normalize_all():
    data_path=f"./data/kimore_kfold.pkl"
    raw_data=joblib.load(data_path)

    processor = PreNormalize3D()
    exs = {}
    for ex in range(1, 6):
        datas = {}
        for split in ['train_data', 'test_data']:
            print(f"Normalizing ex{ex} {split}...")
            datas[split] = normalize_kfold(raw_data[ex][split], processor)
        datas['train_labels'] = raw_data[ex]['train_labels']
        datas['test_labels'] = raw_data[ex]['test_labels']
        datas['selected_subjects'] = raw_data[ex]['selected_subjects']
        datas['train_subjects'] = raw_data[ex]['train_subjects']
        exs[ex] = datas
    joblib.dump(exs, f"./data/kimore_kfold_norm.pkl")

if __name__ == "__main__":
    if os.path.exists("./data/kimore_kfold.pkl"):
        print("kimore_kfold.pkl already exists.")
    else:
        if not os.path.exists("./data/kimore_raw.pkl"):
            print("kimore_raw.pkl not found. Extracting raw data...")
            get_raw_data()
        loader=Data_Loader(path="./data/kimore_raw.pkl", loocv=False)
        ex_kfold={}
        for i in range(1, 6):
            train_data, train_labels, test_data, test_labels, selected_subjects, train_subjects = loader.getkdata(exercise=i, k=5)
            ex_kfold[i]={
                "train_data": train_data,
                "train_labels": train_labels,
                "test_data": test_data,
                "test_labels": test_labels,
                "selected_subjects": selected_subjects,
                "train_subjects": train_subjects
            }
            print(f"Exercise {i} k-fold data loaded")
            print(f"shape: {train_data[0].shape}, labels shape: {train_labels[0].shape}")
        joblib.dump(ex_kfold, "./data/kimore_kfold.pkl")
    normalize_all()